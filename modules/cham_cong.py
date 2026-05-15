"""Module Chấm công nhân viên — DLW Phase 2 + 4 + 5 + 6 + 7.

Tabs trong "👥 Nhân viên":
  - ⚙️ Cấu hình (Phase 2): 3 sub-tabs Mạng / Lương / Ca làm việc
  - 📅 Lịch làm việc (Phase 2): calendar tuần + CRUD schedule
  - 📊 Bảng công (Phase 4): sessions theo date range + summary per NV
    + ✏️ Sửa công (Phase 5, admin only): dialog edit session với audit history
  - 💰 Tính lương (Phase 6): CRUD kỳ lương / bảng lương / phụ cấp-thưởng
    + 📥 Export Excel (Phase 7, admin only): 2-sheet XLSX bảng công + bảng lương

Phase 8 (POS repo): "Lương của tôi" cho NV xem qua POS dialog.

Refs: PLAN_CHAM_CONG.md sections 7 + 9 + 10 + 11 + 12.
"""
from __future__ import annotations

import streamlit as st
import pandas as pd
from datetime import date, datetime, time, timedelta
from zoneinfo import ZoneInfo

from utils.auth import is_admin, is_ke_toan_or_admin, get_user
from utils.config import ALL_BRANCHES
from utils.db import (
    supabase, log_action, call_rpc,
    load_all_nhan_vien,
    load_shift_templates, load_branch_networks, load_employee_rates,
    count_schedules_using_template, load_schedules_for_week,
)

VN_TZ = ZoneInfo("Asia/Ho_Chi_Minh")
DOW_VN = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]


def module_cham_cong():
    """Entry: menu '👥 Nhân viên' cho admin + kế toán."""
    if not is_ke_toan_or_admin():
        st.error("⛔ Tính năng này dành cho admin và kế toán.")
        return

    st.title("👥 Nhân viên")

    tabs = st.tabs([
        "⚙️ Cấu hình",
        "📅 Lịch làm việc",
        "📊 Bảng công",
        "💰 Tính lương",
    ])
    with tabs[0]:
        _render_config()
    with tabs[1]:
        _render_schedule()
    with tabs[2]:
        _render_bang_cong()
    with tabs[3]:
        _render_payroll()


# ============================================================
# ⚙️ CẤU HÌNH — admin only
# ============================================================

def _render_config():
    if not is_admin():
        st.info("ℹ️ Chỉ admin được cấu hình. Bạn có thể xem ở tab Lịch.")
        return

    st.subheader("Cấu hình hệ thống chấm công")
    sub = st.tabs([
        "🌐 Mạng cửa hàng",
        "💵 Lương nhân viên",
        "⏰ Ca làm việc",
    ])
    with sub[0]:
        _config_branch_networks()
    with sub[1]:
        _config_employee_rates()
    with sub[2]:
        _config_shift_templates()


def _config_branch_networks():
    """Set ip_prefixes per CN."""
    st.markdown("##### 🌐 Cấu hình IP whitelist cho từng chi nhánh")
    st.caption(
        "Lấy IP public từ `whatismyip.com` khi đang dùng Wi-Fi cửa hàng. "
        "Nhập 3 số đầu + dấu chấm, vd `113.161.74.`. Mỗi prefix 1 dòng."
    )

    networks = load_branch_networks()
    user = get_user()

    for cn in ALL_BRANCHES:
        with st.container(border=True):
            st.markdown(f"**{cn}**")
            current = networks.get(cn, [])
            current_text = "\n".join(current)

            new_text = st.text_area(
                "IP prefixes (mỗi dòng 1 prefix)",
                value=current_text,
                key=f"cc_net_{cn}",
                height=80,
                placeholder="113.161.74.\n14.225.10.",
                label_visibility="collapsed",
            )

            col1, col2 = st.columns([1, 4])
            with col1:
                if st.button("💾 Lưu", key=f"cc_net_save_{cn}", type="primary"):
                    new_prefixes = [
                        line.strip() for line in new_text.split("\n")
                        if line.strip()
                    ]
                    try:
                        supabase.table("attendance_branch_networks").update({
                            "ip_prefixes": new_prefixes,
                            "updated_by": user["id"],
                            "updated_at": datetime.now(VN_TZ).isoformat(),
                        }).eq("branch_name", cn).execute()
                        log_action("ATT_BRANCH_NET_UPDATE",
                                   f"{cn}: {len(new_prefixes)} prefix(es)")
                        load_branch_networks.clear()
                        st.toast(f"✓ Đã lưu IP cho {cn}", icon="✅")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Lỗi: {e}")
            with col2:
                if current:
                    st.caption(f"Hiện có: {', '.join(current)}")
                else:
                    st.caption("⚠️ Chưa cấu hình → NV không chấm công được")


def _config_employee_rates():
    """Set salary_type + rate per NV."""
    st.markdown("##### 💵 Cấu hình lương theo nhân viên")
    st.caption("Lương theo giờ (hourly) hoặc cố định tháng (monthly_fixed).")

    nv_list = load_all_nhan_vien(include_inactive=False)
    rates = load_employee_rates()
    user = get_user()

    role_filter = st.selectbox(
        "Lọc theo vai trò",
        ["Tất cả", "admin", "ke_toan", "nhan_vien"],
        key="cc_rate_role_filter",
    )
    if role_filter != "Tất cả":
        nv_list = [n for n in nv_list if n.get("role") == role_filter]

    if not nv_list:
        st.info("Không có NV phù hợp.")
        return

    for nv in nv_list:
        nv_id = nv["id"]
        nv_name = nv["ho_ten"]
        nv_role = nv.get("role", "")
        rate = rates.get(nv_id)

        if rate:
            stype = rate.get("salary_type", "")
            if stype == "hourly":
                badge = f"⏰ {int(rate.get('hourly_rate') or 0):,}đ/giờ"
            else:
                badge = f"📅 {int(rate.get('monthly_fixed') or 0):,}đ/tháng"
        else:
            badge = "⚠️ Chưa cấu hình"

        with st.expander(f"**{nv_name}** ({nv_role}) — {badge}"):
            stype_now = (rate or {}).get("salary_type", "hourly")
            stype = st.radio(
                "Loại lương",
                ["hourly", "monthly_fixed"],
                index=0 if stype_now == "hourly" else 1,
                format_func=lambda x: "⏰ Theo giờ" if x == "hourly" else "📅 Cố định tháng",
                key=f"cc_rate_type_{nv_id}",
                horizontal=True,
            )

            if stype == "hourly":
                hourly_val = st.number_input(
                    "Đơn giá (đ/giờ)",
                    min_value=0, step=1000,
                    value=int((rate or {}).get("hourly_rate") or 0),
                    key=f"cc_rate_hourly_{nv_id}",
                )
                fixed_val = None
            else:
                fixed_val = st.number_input(
                    "Lương cố định (đ/tháng)",
                    min_value=0, step=100000,
                    value=int((rate or {}).get("monthly_fixed") or 0),
                    key=f"cc_rate_fixed_{nv_id}",
                )
                hourly_val = None

            if st.button("💾 Lưu", key=f"cc_rate_save_{nv_id}", type="primary"):
                if stype == "hourly" and (not hourly_val or hourly_val <= 0):
                    st.error("Đơn giá phải > 0")
                    return
                if stype == "monthly_fixed" and (not fixed_val or fixed_val <= 0):
                    st.error("Lương cố định phải > 0")
                    return
                try:
                    payload = {
                        "nhan_vien_id": nv_id,
                        "salary_type": stype,
                        "hourly_rate": hourly_val if stype == "hourly" else None,
                        "monthly_fixed": fixed_val if stype == "monthly_fixed" else None,
                        "updated_by": user["id"],
                        "updated_at": datetime.now(VN_TZ).isoformat(),
                    }
                    if rate:
                        supabase.table("attendance_employee_rates") \
                            .update(payload).eq("nhan_vien_id", nv_id).execute()
                    else:
                        supabase.table("attendance_employee_rates") \
                            .insert(payload).execute()
                    log_action("ATT_RATE_UPDATE",
                               f"NV {nv_name}: {stype}={hourly_val or fixed_val}")
                    load_employee_rates.clear()
                    st.toast(f"✓ Đã lưu lương {nv_name}", icon="✅")
                    st.rerun()
                except Exception as e:
                    st.error(f"Lỗi: {e}")


def _config_shift_templates():
    """CRUD shift_templates với lock fields nếu có schedule ref. D17."""
    st.markdown("##### ⏰ Ca làm việc")
    st.caption(
        "Admin tạo/sửa ca. Ca đã có schedule ref → không sửa được giờ/CN "
        "(tạo ca mới + disable ca cũ nếu cần đổi)."
    )

    col_filter, col_show, col_add = st.columns([2, 2, 1])
    with col_filter:
        cn_filter = st.selectbox(
            "Chi nhánh",
            ["Tất cả"] + ALL_BRANCHES,
            key="cc_shift_cn_filter",
        )
    with col_show:
        show_inactive = st.checkbox(
            "Hiện cả ca đã ẩn", value=False, key="cc_shift_show_inactive"
        )
    with col_add:
        st.write("")
        if st.button("➕ Thêm ca", key="cc_shift_add_btn", use_container_width=True):
            _shift_add_dialog()

    templates = load_shift_templates(include_inactive=show_inactive)
    if cn_filter != "Tất cả":
        templates = [t for t in templates if t.get("branch_name") == cn_filter]

    if not templates:
        st.info("Chưa có ca làm việc nào.")
        return

    for t in templates:
        ref_count = count_schedules_using_template(t["id"])
        active = t.get("active", True)
        is_tech = t.get("is_technician", False)
        loai_lbl = "KTV" if is_tech else "NV"
        active_badge = "✅" if active else "⛔"

        with st.container(border=True):
            col_info, col_meta, col_act = st.columns([3, 2, 1])
            with col_info:
                st.markdown(
                    f"**{t['code']}** — {t['label']}  \n"
                    f"📍 {t['branch_name']} · ⏰ {t['start_time'][:5]}–{t['end_time'][:5]} "
                    f"({t['default_hours']}h) · {loai_lbl}"
                )
            with col_meta:
                st.caption(
                    f"{active_badge} {'Active' if active else 'Đã ẩn'} · "
                    f"📋 {ref_count} schedule ref"
                )
            with col_act:
                if st.button("✏️ Sửa", key=f"cc_shift_edit_{t['id']}",
                             use_container_width=True):
                    _shift_edit_dialog(t, ref_count)


@st.dialog("➕ Thêm ca làm việc")
def _shift_add_dialog():
    code = st.text_input(
        "Code (snake_case, unique)", placeholder="vd: tech_coop",
        key="cc_shift_new_code"
    )
    branch = st.selectbox("Chi nhánh", ALL_BRANCHES, key="cc_shift_new_branch")
    label = st.text_input(
        "Label hiển thị", placeholder="vd: Ca KTV Coop",
        key="cc_shift_new_label"
    )

    col1, col2 = st.columns(2)
    with col1:
        start_t = st.time_input("Giờ bắt đầu", value=time(7, 0),
                                key="cc_shift_new_start")
    with col2:
        end_t = st.time_input("Giờ kết thúc", value=time(14, 0),
                              key="cc_shift_new_end")

    is_tech = st.checkbox("Ca KTV (kỹ thuật viên)", key="cc_shift_new_tech")

    s_min = start_t.hour * 60 + start_t.minute
    e_min = end_t.hour * 60 + end_t.minute
    if e_min > s_min:
        default_hours = round((e_min - s_min) / 60)
        st.caption(f"⏱ Tự động: **{default_hours} giờ**")
        valid_time = True
    else:
        default_hours = 0
        st.error("Giờ kết thúc phải sau giờ bắt đầu (v1 chưa hỗ trợ ca qua đêm)")
        valid_time = False

    if st.button("💾 Tạo ca", type="primary", use_container_width=True,
                 disabled=not (code.strip() and label.strip() and valid_time)):
        code_clean = code.strip()
        check = supabase.table("shift_templates").select("id") \
            .eq("code", code_clean).execute()
        if check.data:
            st.error(f"Code '{code_clean}' đã tồn tại. Chọn code khác.")
            return
        try:
            supabase.table("shift_templates").insert({
                "code": code_clean,
                "branch_name": branch,
                "label": label.strip(),
                "start_time": start_t.strftime("%H:%M:%S"),
                "end_time": end_t.strftime("%H:%M:%S"),
                "default_hours": default_hours,
                "is_technician": is_tech,
                "active": True,
            }).execute()
            log_action("ATT_SHIFT_TEMPLATE_CREATE",
                       f"{code_clean} ({branch} {start_t}-{end_t})")
            load_shift_templates.clear()
            st.toast(f"✓ Đã tạo ca {code_clean}", icon="✅")
            st.rerun()
        except Exception as e:
            st.error(f"Lỗi: {e}")


@st.dialog("✏️ Sửa ca làm việc")
def _shift_edit_dialog(template: dict, ref_count: int):
    t_id = template["id"]
    has_refs = ref_count > 0

    st.markdown(f"**Code:** `{template['code']}` (không sửa được)")
    if has_refs:
        st.warning(
            f"⚠️ Ca này đã có **{ref_count} schedule ref**. "
            "Không thể đổi giờ/chi nhánh/loại. Tạo ca mới + disable ca này nếu cần đổi."
        )

    label = st.text_input("Label", value=template["label"],
                          key=f"cc_shift_edit_label_{t_id}")

    branch_idx = (ALL_BRANCHES.index(template["branch_name"])
                  if template.get("branch_name") in ALL_BRANCHES else 0)
    branch = st.selectbox(
        "Chi nhánh", ALL_BRANCHES,
        index=branch_idx,
        disabled=has_refs,
        key=f"cc_shift_edit_branch_{t_id}",
    )

    col1, col2 = st.columns(2)
    with col1:
        start_t = st.time_input(
            "Giờ bắt đầu",
            value=datetime.strptime(template["start_time"][:5], "%H:%M").time(),
            disabled=has_refs,
            key=f"cc_shift_edit_start_{t_id}",
        )
    with col2:
        end_t = st.time_input(
            "Giờ kết thúc",
            value=datetime.strptime(template["end_time"][:5], "%H:%M").time(),
            disabled=has_refs,
            key=f"cc_shift_edit_end_{t_id}",
        )

    is_tech = st.checkbox(
        "Ca KTV", value=template.get("is_technician", False),
        disabled=has_refs,
        key=f"cc_shift_edit_tech_{t_id}",
    )
    active = st.checkbox(
        "Active (uncheck = ẩn ca khỏi UI xếp lịch)",
        value=template.get("active", True),
        key=f"cc_shift_edit_active_{t_id}",
    )

    s_min = start_t.hour * 60 + start_t.minute
    e_min = end_t.hour * 60 + end_t.minute
    default_hours = (round((e_min - s_min) / 60) if e_min > s_min
                     else template.get("default_hours", 0))

    if st.button("💾 Lưu", type="primary", use_container_width=True,
                 disabled=not label.strip(),
                 key=f"cc_shift_edit_save_{t_id}"):
        try:
            payload = {
                "label": label.strip(),
                "active": active,
            }
            if not has_refs:
                payload.update({
                    "branch_name": branch,
                    "start_time": start_t.strftime("%H:%M:%S"),
                    "end_time": end_t.strftime("%H:%M:%S"),
                    "default_hours": default_hours,
                    "is_technician": is_tech,
                })
            supabase.table("shift_templates").update(payload) \
                .eq("id", t_id).execute()
            log_action("ATT_SHIFT_TEMPLATE_EDIT",
                       f"id={t_id} active={active} label={label!r}")
            load_shift_templates.clear()
            st.toast(f"✓ Đã cập nhật ca {template['code']}", icon="✅")
            st.rerun()
        except Exception as e:
            st.error(f"Lỗi: {e}")


# ============================================================
# 📅 LỊCH LÀM VIỆC
# ============================================================

def _get_week_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _render_schedule():
    today = datetime.now(VN_TZ).date()

    col_date, col_branch, col_copy = st.columns([2, 2, 2])
    with col_date:
        anchor_date = st.date_input(
            "Chọn ngày (lấy tuần chứa)",
            value=today,
            key="cc_sched_anchor",
        )
    monday = _get_week_monday(anchor_date)
    sunday = monday + timedelta(days=6)

    with col_branch:
        cn_filter = st.selectbox(
            "Chi nhánh",
            ["Tất cả"] + ALL_BRANCHES,
            key="cc_sched_cn_filter",
        )
    with col_copy:
        st.write("")
        if is_admin():
            if st.button("📋 Copy tuần trước", key="cc_sched_copy_btn",
                         use_container_width=True):
                _copy_previous_week(
                    monday,
                    branch=None if cn_filter == "Tất cả" else cn_filter,
                )

    st.caption(
        f"📅 Tuần: **{monday.strftime('%d/%m')} → {sunday.strftime('%d/%m/%Y')}**"
    )

    branch_arg = None if cn_filter == "Tất cả" else cn_filter
    schedules = load_schedules_for_week(monday, branch=branch_arg)

    by_date: dict[date, list] = {monday + timedelta(days=i): [] for i in range(7)}
    for s in schedules:
        try:
            wd = datetime.strptime(s["work_date"], "%Y-%m-%d").date()
            if wd in by_date:
                by_date[wd].append(s)
        except Exception:
            continue

    cols = st.columns(7)
    for i, col in enumerate(cols):
        d = monday + timedelta(days=i)
        is_today = (d == today)
        with col:
            day_lbl = f"{DOW_VN[i]} {d.day}/{d.month}"
            if is_today:
                st.markdown(
                    f"<div style='font-weight:700;color:#e63946;'>{day_lbl} (hôm nay)</div>",
                    unsafe_allow_html=True
                )
            else:
                st.markdown(
                    f"<div style='font-weight:600;'>{day_lbl}</div>",
                    unsafe_allow_html=True
                )

            day_scheds = sorted(
                by_date.get(d, []),
                key=lambda s: s.get("scheduled_start_at") or ""
            )
            for s in day_scheds:
                _render_schedule_chip(s)

            if is_admin():
                if st.button("➕", key=f"cc_sched_add_{d.isoformat()}",
                             use_container_width=True,
                             help=f"Thêm lịch {d.strftime('%d/%m')}"):
                    _schedule_add_dialog(d)


def _render_schedule_chip(s: dict):
    tmpl = s.get("shift_templates") or {}
    nv = s.get("nhan_vien") or {}
    nv_name = nv.get("ho_ten", "?")
    status = s.get("status", "scheduled")

    label = tmpl.get("label", "?")

    status_lbl = {
        "scheduled": "",
        "leave_paid": " 🌴",
        "leave_unpaid": " ⚪",
    }.get(status, "")

    btn_label = f"{nv_name} · {label}{status_lbl}"

    if st.button(btn_label, key=f"cc_chip_{s['id']}",
                 use_container_width=True,
                 help=f"{tmpl.get('branch_name','')} · {tmpl.get('start_time','')[:5]}–{tmpl.get('end_time','')[:5]}"):
        _schedule_edit_dialog(s)


@st.dialog("➕ Thêm lịch làm việc")
def _schedule_add_dialog(work_date: date):
    user = get_user()
    nv_list = load_all_nhan_vien(include_inactive=False)
    templates = load_shift_templates(include_inactive=False)

    if not templates:
        st.error("Chưa có ca làm việc nào active. Tạo ca trước trong Cấu hình > Ca làm việc.")
        return

    st.markdown(f"**Ngày:** {work_date.strftime('%d/%m/%Y')}")

    nv_options = {f"{n['ho_ten']} ({n.get('role','')})": n["id"] for n in nv_list}
    nv_label = st.selectbox("Nhân viên", list(nv_options.keys()),
                            key=f"cc_sched_new_nv_{work_date.isoformat()}")
    nv_id = nv_options[nv_label]

    tmpl_options = {
        f"{t['branch_name']} — {t['label']} ({t['start_time'][:5]}-{t['end_time'][:5]})": t
        for t in templates
    }
    tmpl_label = st.selectbox("Ca làm việc", list(tmpl_options.keys()),
                              key=f"cc_sched_new_tmpl_{work_date.isoformat()}")
    tmpl = tmpl_options[tmpl_label]

    status = st.selectbox(
        "Trạng thái",
        ["scheduled", "leave_paid", "leave_unpaid"],
        format_func=lambda s: {
            "scheduled": "✅ Xếp lịch",
            "leave_paid": "🌴 Nghỉ phép có lương",
            "leave_unpaid": "⚪ Nghỉ không lương",
        }[s],
        key=f"cc_sched_new_status_{work_date.isoformat()}",
    )
    note = st.text_area("Ghi chú (optional)", height=60,
                        key=f"cc_sched_new_note_{work_date.isoformat()}")

    if st.button("💾 Tạo lịch", type="primary", use_container_width=True,
                 key=f"cc_sched_new_save_{work_date.isoformat()}"):
        start_dt = datetime.combine(
            work_date,
            datetime.strptime(tmpl["start_time"][:5], "%H:%M").time(),
            tzinfo=VN_TZ,
        )
        end_dt = datetime.combine(
            work_date,
            datetime.strptime(tmpl["end_time"][:5], "%H:%M").time(),
            tzinfo=VN_TZ,
        )

        if status == "scheduled":
            err = _check_overlap(nv_id, start_dt, end_dt, exclude_schedule_id=None)
            if err:
                st.error(err)
                return

        try:
            supabase.table("attendance_work_schedules").insert({
                "nhan_vien_id": nv_id,
                "work_date": work_date.isoformat(),
                "shift_template_id": tmpl["id"],
                "scheduled_start_at": start_dt.isoformat(),
                "scheduled_end_at": end_dt.isoformat(),
                "status": status,
                "note": note or None,
                "created_by": user["id"],
            }).execute()
            log_action("ATT_SCHEDULE_CREATE",
                       f"NV {nv_id} {work_date} ca {tmpl['code']} status={status}")
            st.toast("✓ Đã tạo lịch", icon="✅")
            st.rerun()
        except Exception as e:
            msg = str(e)
            if "no_overlap_per_nv" in msg or "exclusion" in msg.lower():
                st.error("Lịch overlap với ca khác của NV này. Kiểm tra lại giờ.")
            elif "unique" in msg.lower() or "duplicate" in msg.lower():
                st.error("NV đã có ca này trong ngày.")
            else:
                st.error(f"Lỗi: {e}")


@st.dialog("✏️ Sửa lịch làm việc")
def _schedule_edit_dialog(schedule: dict):
    tmpl = schedule.get("shift_templates") or {}
    nv = schedule.get("nhan_vien") or {}
    s_id = schedule["id"]

    st.markdown(
        f"**NV:** {nv.get('ho_ten','?')}  \n"
        f"**Ngày:** {schedule.get('work_date','?')}  \n"
        f"**Ca:** {tmpl.get('label','?')} "
        f"({tmpl.get('start_time','')[:5]}-{tmpl.get('end_time','')[:5]} · "
        f"{tmpl.get('branch_name','')})"
    )

    statuses = ["scheduled", "cancelled", "leave_paid", "leave_unpaid"]
    current_status = schedule.get("status", "scheduled")
    if current_status not in statuses:
        current_status = "scheduled"

    status = st.selectbox(
        "Trạng thái",
        statuses,
        index=statuses.index(current_status),
        format_func=lambda s: {
            "scheduled": "✅ Xếp lịch",
            "cancelled": "❌ Hủy lịch",
            "leave_paid": "🌴 Nghỉ phép có lương",
            "leave_unpaid": "⚪ Nghỉ không lương",
        }[s],
        key=f"cc_sched_edit_status_{s_id}",
        disabled=not is_admin(),
    )
    note = st.text_area(
        "Ghi chú",
        value=schedule.get("note") or "",
        height=60,
        key=f"cc_sched_edit_note_{s_id}",
        disabled=not is_admin(),
    )

    if is_admin():
        col1, col2 = st.columns(2)
        with col1:
            if st.button("💾 Lưu", type="primary", use_container_width=True,
                         key=f"cc_sched_edit_save_{s_id}"):
                try:
                    supabase.table("attendance_work_schedules").update({
                        "status": status,
                        "note": note or None,
                        "updated_at": datetime.now(VN_TZ).isoformat(),
                    }).eq("id", s_id).execute()
                    log_action("ATT_SCHEDULE_EDIT", f"id={s_id} status={status}")
                    st.toast("✓ Đã cập nhật", icon="✅")
                    st.rerun()
                except Exception as e:
                    st.error(f"Lỗi: {e}")
        with col2:
            if st.button("🗑 Xóa lịch", use_container_width=True,
                         key=f"cc_sched_edit_del_{s_id}"):
                try:
                    supabase.table("attendance_work_schedules") \
                        .delete().eq("id", s_id).execute()
                    log_action("ATT_SCHEDULE_DELETE", f"id={s_id}")
                    st.toast("✓ Đã xóa lịch", icon="🗑")
                    st.rerun()
                except Exception as e:
                    st.error(f"Lỗi: {e}")
    else:
        st.caption("ℹ️ Chỉ admin được sửa/xóa lịch.")


def _check_overlap(nv_id: int, start_dt: datetime, end_dt: datetime,
                   exclude_schedule_id: int | None = None) -> str | None:
    """Pre-check overlap trong Python — error rõ ràng trước khi DB constraint reject."""
    lo = (start_dt - timedelta(hours=24)).isoformat()
    hi = (end_dt + timedelta(hours=24)).isoformat()
    res = supabase.table("attendance_work_schedules").select(
        "id, scheduled_start_at, scheduled_end_at, shift_templates(label)"
    ).eq("nhan_vien_id", nv_id).eq("status", "scheduled") \
     .gte("scheduled_start_at", lo).lte("scheduled_start_at", hi).execute()

    for r in res.data or []:
        if exclude_schedule_id and r["id"] == exclude_schedule_id:
            continue
        try:
            rs = datetime.fromisoformat(r["scheduled_start_at"])
            re_ = datetime.fromisoformat(r["scheduled_end_at"])
        except Exception:
            continue
        if start_dt < re_ and end_dt > rs:
            lbl = (r.get("shift_templates") or {}).get("label", "?")
            return (
                f"NV đã có ca **{lbl}** từ "
                f"{rs.astimezone(VN_TZ).strftime('%d/%m %H:%M')} đến "
                f"{re_.astimezone(VN_TZ).strftime('%H:%M')} — overlap"
            )
    return None


def _copy_previous_week(target_monday: date, branch: str | None = None):
    """Copy scheduled lịch tuần trước → tuần này. Skip cancelled/leave_*/trùng/lỗi."""
    user = get_user()
    prev_monday = target_monday - timedelta(days=7)
    src_schedules = load_schedules_for_week(prev_monday, branch=branch)
    src_schedules = [s for s in src_schedules if s.get("status") == "scheduled"]

    if not src_schedules:
        st.toast("⚠️ Tuần trước không có lịch scheduled nào", icon="⚠️")
        return

    created = 0
    skipped = 0
    for s in src_schedules:
        try:
            wd_old = datetime.strptime(s["work_date"], "%Y-%m-%d").date()
            offset = (wd_old - prev_monday).days
            wd_new = target_monday + timedelta(days=offset)

            tmpl = s.get("shift_templates") or {}
            if not tmpl or not tmpl.get("active", True):
                skipped += 1
                continue

            start_dt = datetime.combine(
                wd_new,
                datetime.strptime(tmpl["start_time"][:5], "%H:%M").time(),
                tzinfo=VN_TZ,
            )
            end_dt = datetime.combine(
                wd_new,
                datetime.strptime(tmpl["end_time"][:5], "%H:%M").time(),
                tzinfo=VN_TZ,
            )

            supabase.table("attendance_work_schedules").insert({
                "nhan_vien_id": s["nhan_vien_id"],
                "work_date": wd_new.isoformat(),
                "shift_template_id": s["shift_template_id"],
                "scheduled_start_at": start_dt.isoformat(),
                "scheduled_end_at": end_dt.isoformat(),
                "status": "scheduled",
                "created_by": user["id"],
            }).execute()
            created += 1
        except Exception:
            skipped += 1

    log_action("ATT_SCHEDULE_COPY_WEEK",
               f"{prev_monday}→{target_monday}: {created} OK, {skipped} skipped")
    st.toast(f"✓ Copy: {created} lịch (skip {skipped} trùng/lỗi)", icon="✅")
    st.rerun()


# ============================================================
# 📊 BẢNG CÔNG (Phase 4)
# ============================================================

_STATUS_LABEL_VN = {
    "open": "🔵 Đang trong ca",
    "completed": "✅ Hoàn thành",
    "auto_closed": "🟠 Auto đóng",
    "absent": "🔴 Vắng",
    "edited": "🟢 Đã sửa",
    "leave_paid": "🌴 Nghỉ phép",
    "leave_unpaid": "⚪ Nghỉ KL",
}


def _render_bang_cong():
    """Bảng công — sessions theo date range, role-aware filter."""
    today = datetime.now(VN_TZ).date()
    first_of_month = today.replace(day=1)

    col_from, col_to, col_branch = st.columns([2, 2, 2])
    with col_from:
        date_from = st.date_input("Từ ngày", value=first_of_month, key="cc_bc_from")
    with col_to:
        date_to = st.date_input("Đến ngày", value=today, key="cc_bc_to")
    with col_branch:
        cn_filter = st.selectbox(
            "Chi nhánh", ["Tất cả"] + ALL_BRANCHES, key="cc_bc_cn"
        )

    if date_from > date_to:
        st.error("Ngày 'Từ' phải <= 'Đến'")
        return
    if (date_to - date_from).days > 92:
        st.warning("⚠️ Range > 3 tháng có thể load chậm.")

    branch_arg = None if cn_filter == "Tất cả" else cn_filter

    # NV filter (role-aware)
    visible_nv_ids = _visible_nv_ids()
    nv_options = load_all_nhan_vien(include_inactive=True)
    if visible_nv_ids is not None:
        nv_options = [n for n in nv_options if n["id"] in visible_nv_ids]

    nv_names_selected = st.multiselect(
        "Nhân viên (rỗng = tất cả NV thuộc quyền xem)",
        options=[n["ho_ten"] for n in nv_options],
        default=[],
        key="cc_bc_nv",
    )

    nv_id_filter: list[int] | None = None
    if nv_names_selected:
        nv_id_filter = [n["id"] for n in nv_options
                        if n["ho_ten"] in nv_names_selected]
    elif visible_nv_ids is not None:
        nv_id_filter = visible_nv_ids

    # Build trigger
    col_build, col_info = st.columns([2, 4])
    with col_build:
        if st.button("🔄 Cập nhật sessions",
                     help="Build sessions từ events thô (idempotent)",
                     key="cc_bc_build_btn"):
            _trigger_build_sessions(date_from, date_to, branch_arg)
            st.rerun()
    with col_info:
        st.caption(
            "💡 Bấm 'Cập nhật sessions' nếu vừa có chấm công mới — "
            "tự động trigger `build_sessions_for_date` cho từng ngày trong range."
        )

    sessions = _load_sessions_for_range(
        date_from, date_to, branch_arg, nv_id_filter
    )

    if not sessions:
        st.info(
            f"📭 Chưa có session nào trong khoảng "
            f"**{date_from.strftime('%d/%m')} → {date_to.strftime('%d/%m/%Y')}**. "
            "Bấm '🔄 Cập nhật sessions' để build từ events thô (nếu NV đã chấm công)."
        )
        return

    df = _sessions_to_df(sessions)
    _render_bang_cong_table(df)

    # ✏️ Sửa session — admin only (Phase 5)
    if is_admin():
        _render_session_edit_picker(sessions)

    st.markdown("---")
    st.markdown("##### 📊 Tổng kết theo nhân viên")
    _render_bang_cong_summary(df)


def _visible_nv_ids() -> list[int] | None:
    """Role-aware filter: admin all, ke_toan ex-admin, NV chỉ mình.
    Return None = no filter (load all)."""
    u = get_user()
    if not u:
        return [-1]  # nothing accessible
    role = u.get("role", "")
    if role == "admin":
        return None
    if role == "ke_toan":
        res = supabase.table("nhan_vien").select("id") \
            .neq("role", "admin").execute()
        return [n["id"] for n in (res.data or [])]
    return [u["id"]]


def _trigger_build_sessions(date_from: date, date_to: date,
                            branch: str | None):
    """Call RPC build_sessions_for_date per day in range. Idempotent.

    Aggregate by_status để hiển thị edited count (sessions admin-edited
    được preserve, không recompute — hotfix Phase 5).
    """
    d = date_from
    total = 0
    errors = 0
    by_status_agg: dict[str, int] = {}
    while d <= date_to:
        try:
            res = call_rpc("build_sessions_for_date", {
                "p_work_date": d.isoformat(),
                "p_chi_nhanh": branch,
            })
            if isinstance(res, dict) and res.get("ok"):
                total += int(res.get("sessions_count", 0) or 0)
                for k, v in (res.get("by_status") or {}).items():
                    by_status_agg[k] = by_status_agg.get(k, 0) + int(v)
            else:
                errors += 1
        except Exception:
            errors += 1
        d += timedelta(days=1)

    if errors:
        st.toast(f"⚠️ Built {total} sessions, {errors} ngày lỗi",
                 icon="⚠️")
        return

    msg = f"✓ Đã cập nhật {total} sessions"
    edited_count = by_status_agg.get("edited", 0)
    if edited_count > 0:
        msg += f" ({edited_count} đã sửa — giữ nguyên không recompute)"
    st.toast(msg, icon="✅")


def _load_sessions_for_range(date_from: date, date_to: date,
                              branch: str | None,
                              nv_id_filter: list[int] | None) -> list[dict]:
    """Load sessions + schedule + template + nv info merged. Sort newest-first."""
    # 1. Load schedules in range (embed template only — nhan_vien có 2 FK)
    q = supabase.table("attendance_work_schedules").select(
        "id, nhan_vien_id, work_date, shift_template_id, status, "
        "shift_templates(label, branch_name, is_technician, default_hours)"
    ).gte("work_date", date_from.isoformat()) \
     .lte("work_date", date_to.isoformat())

    if nv_id_filter is not None:
        if not nv_id_filter:
            return []
        q = q.in_("nhan_vien_id", nv_id_filter)

    sched_rows = q.execute().data or []

    if branch:
        sched_rows = [r for r in sched_rows
                      if (r.get("shift_templates") or {}).get("branch_name") == branch]

    if not sched_rows:
        return []

    sched_map = {r["id"]: r for r in sched_rows}
    sched_ids = list(sched_map.keys())

    # 2. Load sessions cho các schedule_ids (batch 200 phòng URL length)
    sess_rows: list[dict] = []
    for i in range(0, len(sched_ids), 200):
        chunk = sched_ids[i:i + 200]
        res = supabase.table("attendance_sessions").select("*") \
            .in_("schedule_id", chunk).execute()
        sess_rows.extend(res.data or [])

    # 3. Load NV info (ho_ten + role)
    nv_ids_needed = list({r["nhan_vien_id"] for r in sched_rows})
    nv_map: dict[int, dict] = {}
    if nv_ids_needed:
        nv_res = supabase.table("nhan_vien").select("id, ho_ten, role") \
            .in_("id", nv_ids_needed).execute()
        nv_map = {n["id"]: n for n in (nv_res.data or [])}

    # 4. Merge — anchor by schedule (mỗi schedule = 1 row)
    sess_by_sched = {s["schedule_id"]: s for s in sess_rows}
    result = []
    for sched_id, sched in sched_map.items():
        s = sess_by_sched.get(sched_id)
        tmpl = sched.get("shift_templates") or {}
        nv = nv_map.get(sched["nhan_vien_id"], {})
        sched_status = sched.get("status") or "scheduled"

        # Skip cancelled schedules (không tính bảng công)
        if sched_status == "cancelled":
            continue

        # Schedule leave_* không có session → tạo virtual row
        if not s and sched_status in ("leave_paid", "leave_unpaid"):
            result.append({
                "session_id": None,
                "schedule_id": sched_id,
                "nhan_vien_id": sched["nhan_vien_id"],
                "ho_ten": nv.get("ho_ten", "?"),
                "role": nv.get("role", ""),
                "work_date": sched.get("work_date", ""),
                "shift_label": tmpl.get("label", "?"),
                "branch_name": tmpl.get("branch_name", ""),
                "is_technician": tmpl.get("is_technician", False),
                "check_in_at": None,
                "check_out_at": None,
                "is_late": False,
                "late_minutes": 0,
                "worked_minutes": 0,
                "regular_minutes": 0,
                "ot_minutes": 0,
                "status": sched_status,
                "is_auto_checkout": False,
                "note": None,
            })
            continue

        if not s:
            # Chưa build session row → skip (UI hint user bấm Cập nhật)
            continue

        result.append({
            "session_id": s["id"],
            "schedule_id": sched_id,
            "nhan_vien_id": s["nhan_vien_id"],
            "ho_ten": nv.get("ho_ten", "?"),
            "role": nv.get("role", ""),
            "work_date": sched.get("work_date", ""),
            "shift_label": tmpl.get("label", "?"),
            "branch_name": tmpl.get("branch_name", ""),
            "is_technician": tmpl.get("is_technician", False),
            "check_in_at": s.get("check_in_at"),
            "check_out_at": s.get("check_out_at"),
            "is_late": bool(s.get("is_late", False)),
            "late_minutes": int(s.get("late_minutes", 0) or 0),
            "worked_minutes": int(s.get("worked_minutes", 0) or 0),
            "regular_minutes": int(s.get("regular_minutes", 0) or 0),
            "ot_minutes": int(s.get("ot_minutes", 0) or 0),
            "status": s.get("status", "open"),
            "is_auto_checkout": bool(s.get("is_auto_checkout", False)),
            "note": s.get("note"),
        })

    # Sort newest first, then by NV
    result.sort(
        key=lambda r: (r["work_date"], r["shift_label"], r["ho_ten"]),
        reverse=True,
    )
    return result


def _format_time_vn(iso_str: str | None) -> str:
    """ISO timestamptz → 'HH:MM' (VN tz). Empty/None → '—'."""
    if not iso_str:
        return "—"
    try:
        dt = datetime.fromisoformat(iso_str)
        return dt.astimezone(VN_TZ).strftime("%H:%M")
    except Exception:
        return "—"


def _format_minutes(m: int) -> str:
    """123 minutes → '2h 3p'. Negative/0 → '0'."""
    if not m or m <= 0:
        return "0"
    h = m // 60
    p = m % 60
    if h and p:
        return f"{h}h {p}p"
    if h:
        return f"{h}h"
    return f"{p}p"


def _sessions_to_df(sessions: list[dict]) -> pd.DataFrame:
    rows = []
    for s in sessions:
        try:
            wd = datetime.strptime(s["work_date"], "%Y-%m-%d").date()
            ngay_str = wd.strftime("%d/%m")
            dow = DOW_VN[wd.weekday()]
        except Exception:
            ngay_str = s.get("work_date", "?")
            dow = ""

        status = s.get("status", "open")
        status_lbl = _STATUS_LABEL_VN.get(status, status)
        if s.get("is_auto_checkout") and status == "auto_closed":
            status_lbl += " ⚠"

        late_str = f"{s['late_minutes']}p" if s.get("is_late") else "—"
        branch_short = (s.get("branch_name") or "")[:10]

        rows.append({
            "Ngày": ngay_str,
            "Thứ": dow,
            "NV": s["ho_ten"],
            "Ca": f"{s['shift_label']} ({branch_short})",
            "Vào": _format_time_vn(s.get("check_in_at")),
            "Ra": _format_time_vn(s.get("check_out_at")),
            "Worked": _format_minutes(s.get("worked_minutes", 0)),
            "OT": _format_minutes(s.get("ot_minutes", 0)) if s.get("ot_minutes") else "—",
            "Late": late_str,
            "Status": status_lbl,
            # Hidden cols cho summary
            "_nv_id": s["nhan_vien_id"],
            "_worked_minutes": s.get("worked_minutes", 0),
            "_ot_minutes": s.get("ot_minutes", 0),
            "_regular_minutes": s.get("regular_minutes", 0),
            "_is_late": bool(s.get("is_late", False)),
            "_status": status,
        })
    return pd.DataFrame(rows)


def _render_bang_cong_table(df: pd.DataFrame):
    if df.empty:
        st.info("Không có session nào.")
        return
    display_cols = ["Ngày", "Thứ", "NV", "Ca", "Vào", "Ra",
                    "Worked", "OT", "Late", "Status"]
    st.dataframe(
        df[display_cols],
        hide_index=True,
        use_container_width=True,
        height=min(36 * (len(df) + 1) + 4, 600),
    )


def _render_bang_cong_summary(df: pd.DataFrame):
    if df.empty:
        return
    summary = df.groupby("NV").agg(
        sessions=("Ngày", "count"),
        regular_minutes=("_regular_minutes", "sum"),
        ot_minutes=("_ot_minutes", "sum"),
        worked_minutes=("_worked_minutes", "sum"),
        late_count=("_is_late", "sum"),
    ).reset_index()

    summary["Giờ regular"] = summary["regular_minutes"].apply(_format_minutes)
    summary["Giờ OT"] = summary["ot_minutes"].apply(_format_minutes)
    summary["Tổng worked"] = summary["worked_minutes"].apply(_format_minutes)
    summary["Late"] = summary["late_count"].astype(int)
    summary = summary.rename(columns={"sessions": "Sessions"})

    st.dataframe(
        summary[["NV", "Sessions", "Giờ regular", "Giờ OT",
                 "Tổng worked", "Late"]],
        hide_index=True,
        use_container_width=True,
    )

    # Overall metrics
    total_sessions = int(summary["Sessions"].sum())
    total_worked = int(summary["worked_minutes"].sum())
    total_ot = int(summary["ot_minutes"].sum())
    total_late = int(summary["late_count"].sum())

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Sessions", total_sessions)
    col2.metric("Tổng worked", _format_minutes(total_worked))
    col3.metric("Tổng OT", _format_minutes(total_ot))
    col4.metric("Late count", total_late)


# ============================================================
# ✏️ SỬA CÔNG (Phase 5) — admin only, audit qua admin_edit_history
# ============================================================

def _render_session_edit_picker(sessions: list[dict]):
    """Sub-section dưới bảng công: chọn session để mở dialog sửa."""
    st.markdown("---")
    st.markdown("##### ✏️ Sửa công")
    st.caption(
        "Sửa giờ vào/ra + ghi chú cho 1 session. Mọi thay đổi được log vào "
        "`admin_edit_history` với lý do bắt buộc."
    )

    editable = [s for s in sessions if s.get("session_id")]
    if not editable:
        st.info(
            "Không có session nào để sửa. Schedule chưa build session "
            "(bấm '🔄 Cập nhật sessions') hoặc chỉ có row leave_* (chưa có session)."
        )
        return

    labels = []
    for s in editable:
        try:
            wd = datetime.strptime(s["work_date"], "%Y-%m-%d").date()
            ngay = wd.strftime("%d/%m")
        except Exception:
            ngay = s.get("work_date", "?")
        labels.append(
            f"#{s['session_id']} · {ngay} · {s['ho_ten']} · "
            f"{s['shift_label']} ({(s.get('branch_name') or '')[:10]}) · "
            f"{_STATUS_LABEL_VN.get(s.get('status',''), s.get('status',''))}"
        )

    col_pick, col_edit = st.columns([4, 1])
    with col_pick:
        picked_label = st.selectbox(
            "Chọn session cần sửa",
            options=labels,
            key="cc_bc_edit_pick",
            label_visibility="collapsed",
        )
    with col_edit:
        if st.button("✏️ Sửa", use_container_width=True, key="cc_bc_edit_btn"):
            idx = labels.index(picked_label)
            _session_edit_dialog(editable[idx])


@st.dialog("✏️ Sửa session chấm công")
def _session_edit_dialog(session_data: dict):
    """Sửa check_in_at, check_out_at, note của 1 session.

    Pattern: RPC update_session_admin snapshot vào admin_edit_history.
    BẮT BUỘC: lý do + gõ "XÁC NHẬN" trước khi Save.
    """
    if not is_admin():
        st.error("⛔ Chỉ admin được sửa session.")
        return

    s_id = session_data.get("session_id")
    sched_id = session_data.get("schedule_id")
    if not s_id or not sched_id:
        st.error("Session không hợp lệ.")
        return

    # Read-only header
    status_lbl = _STATUS_LABEL_VN.get(
        session_data.get("status", ""), session_data.get("status", "")
    )
    st.markdown(
        f"**Session:** `#{s_id}`  \n"
        f"**NV:** {session_data['ho_ten']}  \n"
        f"**Ngày:** {session_data['work_date']}  \n"
        f"**Ca:** {session_data['shift_label']} ({session_data['branch_name']})  \n"
        f"**Status hiện tại:** {status_lbl}"
    )

    # Load scheduled bounds để gợi ý input + recompute reference
    sched_res = supabase.table("attendance_work_schedules").select(
        "scheduled_start_at, scheduled_end_at"
    ).eq("id", sched_id).limit(1).execute()
    sched = (sched_res.data or [{}])[0]
    sched_start_iso = sched.get("scheduled_start_at")
    sched_end_iso = sched.get("scheduled_end_at")

    sched_start_dt = (datetime.fromisoformat(sched_start_iso).astimezone(VN_TZ)
                      if sched_start_iso else None)
    sched_end_dt = (datetime.fromisoformat(sched_end_iso).astimezone(VN_TZ)
                    if sched_end_iso else None)

    if sched_start_dt and sched_end_dt:
        st.caption(
            f"📅 Ca scheduled: **{sched_start_dt.strftime('%d/%m %H:%M')}** "
            f"→ **{sched_end_dt.strftime('%H:%M')}**"
        )

    st.markdown("---")

    # Editable inputs — preserve current values, fallback to scheduled
    curr_in_iso = session_data.get("check_in_at")
    curr_out_iso = session_data.get("check_out_at")
    curr_note = session_data.get("note") or ""

    if curr_in_iso:
        curr_in_dt = datetime.fromisoformat(curr_in_iso).astimezone(VN_TZ)
    elif sched_start_dt:
        curr_in_dt = sched_start_dt
    else:
        curr_in_dt = datetime.now(VN_TZ).replace(microsecond=0)

    if curr_out_iso:
        curr_out_dt = datetime.fromisoformat(curr_out_iso).astimezone(VN_TZ)
    elif sched_end_dt:
        curr_out_dt = sched_end_dt
    else:
        curr_out_dt = curr_in_dt + timedelta(hours=7)

    st.markdown("**Giờ vào (check_in_at):**")
    col_in_d, col_in_t = st.columns([1, 1])
    with col_in_d:
        new_in_date = st.date_input(
            "Ngày vào",
            value=curr_in_dt.date(),
            key=f"cc_edit_in_d_{s_id}",
            label_visibility="collapsed",
        )
    with col_in_t:
        new_in_time = st.time_input(
            "Giờ vào",
            value=curr_in_dt.time().replace(microsecond=0),
            key=f"cc_edit_in_t_{s_id}",
            label_visibility="collapsed",
            step=60,
        )

    st.markdown("**Giờ ra (check_out_at):**")
    col_out_d, col_out_t = st.columns([1, 1])
    with col_out_d:
        new_out_date = st.date_input(
            "Ngày ra",
            value=curr_out_dt.date(),
            key=f"cc_edit_out_d_{s_id}",
            label_visibility="collapsed",
        )
    with col_out_t:
        new_out_time = st.time_input(
            "Giờ ra",
            value=curr_out_dt.time().replace(microsecond=0),
            key=f"cc_edit_out_t_{s_id}",
            label_visibility="collapsed",
            step=60,
        )

    new_note = st.text_area(
        "Ghi chú session",
        value=curr_note,
        key=f"cc_edit_note_{s_id}",
        height=68,
    )

    new_in_dt = datetime.combine(new_in_date, new_in_time, tzinfo=VN_TZ)
    new_out_dt = datetime.combine(new_out_date, new_out_time, tzinfo=VN_TZ)

    if new_out_dt <= new_in_dt:
        st.error("⛔ Giờ ra phải sau giờ vào.")
        return

    # Compute diff preview
    def _fmt_dt_diff(dt):
        return dt.astimezone(VN_TZ).strftime("%d/%m %H:%M") if dt else "—"

    diffs = []
    if curr_in_iso:
        old_in_dt = datetime.fromisoformat(curr_in_iso)
        if abs((new_in_dt - old_in_dt).total_seconds()) > 30:
            diffs.append(("check_in_at",
                          _fmt_dt_diff(old_in_dt), _fmt_dt_diff(new_in_dt)))
    else:
        diffs.append(("check_in_at", "—", _fmt_dt_diff(new_in_dt)))

    if curr_out_iso:
        old_out_dt = datetime.fromisoformat(curr_out_iso)
        if abs((new_out_dt - old_out_dt).total_seconds()) > 30:
            diffs.append(("check_out_at",
                          _fmt_dt_diff(old_out_dt), _fmt_dt_diff(new_out_dt)))
    else:
        diffs.append(("check_out_at", "—", _fmt_dt_diff(new_out_dt)))

    if (new_note or "") != (curr_note or ""):
        diffs.append(("note",
                      curr_note or "(rỗng)", new_note or "(rỗng)"))

    if not diffs:
        st.info("ℹ️ Chưa có thay đổi nào.")
    else:
        with st.container(border=True):
            st.markdown("**📝 Preview thay đổi:**")
            for field, old_v, new_v in diffs:
                st.caption(f"• `{field}`: ~~{old_v}~~ → **{new_v}**")

    st.markdown("---")

    # Lý do (BẮT BUỘC) + confirm typing
    reason = st.text_area(
        "**Lý do sửa** (bắt buộc)",
        key=f"cc_edit_reason_{s_id}",
        height=68,
        placeholder="vd: Máy chấm công lỗi, NV quên chấm ra...",
    )

    confirm_text = st.text_input(
        "Gõ `XÁC NHẬN` để bật nút Save",
        key=f"cc_edit_confirm_{s_id}",
        placeholder="XÁC NHẬN",
    )

    has_changes = bool(diffs)
    reason_ok = bool(reason.strip())
    confirm_ok = confirm_text.strip().upper() == "XÁC NHẬN"
    can_save = has_changes and reason_ok and confirm_ok

    if not has_changes:
        st.caption("💡 Không có thay đổi để lưu.")
    elif not reason_ok:
        st.caption("💡 Nhập lý do để bật nút Save.")
    elif not confirm_ok:
        st.caption("💡 Gõ chính xác `XÁC NHẬN` (in hoa) để bật nút Save.")

    if st.button(
        "💾 Lưu thay đổi",
        type="primary",
        use_container_width=True,
        disabled=not can_save,
        key=f"cc_edit_save_{s_id}",
    ):
        try:
            user = get_user()
            res = call_rpc("update_session_admin", {
                "p_session_id": s_id,
                "p_check_in_at": new_in_dt.isoformat(),
                "p_check_out_at": new_out_dt.isoformat(),
                "p_note": new_note or None,
                "p_reason": reason.strip(),
                "p_admin_id": user["id"],
            })
            if isinstance(res, dict) and res.get("ok"):
                fields = res.get("fields_changed", []) or []
                st.toast(
                    f"✓ Đã sửa session — {len(fields)} field(s) changed",
                    icon="✅",
                )
                st.rerun()
            else:
                err = (res or {}).get("error", "Lỗi RPC")
                st.error(f"❌ {err}")
        except Exception as e:
            st.error(f"Lỗi: {e}")

    # History viewer
    st.markdown("---")
    with st.expander("📜 Lịch sử chỉnh sửa", expanded=False):
        _render_session_edit_history(s_id)


def _render_session_edit_history(session_id: int):
    """Load admin_edit_history cho 1 session, render dạng list."""
    try:
        res = supabase.table("admin_edit_history").select(
            "id, snapshot_before, snapshot_after, fields_changed, "
            "edited_by_name, edit_reason, edited_at"
        ).eq("table_name", "attendance_sessions") \
         .eq("record_id", str(session_id)) \
         .order("edited_at", desc=True) \
         .execute()
        rows = res.data or []
    except Exception as e:
        st.error(f"Lỗi load lịch sử: {e}")
        return

    if not rows:
        st.caption("Chưa có lịch sử sửa cho session này.")
        return

    for r in rows:
        edited_at = r.get("edited_at", "")
        try:
            edited_at_dt = datetime.fromisoformat(edited_at).astimezone(VN_TZ)
            edited_at_str = edited_at_dt.strftime("%d/%m/%Y %H:%M")
        except Exception:
            edited_at_str = edited_at

        fields = r.get("fields_changed") or []
        reason = r.get("edit_reason") or "(không ghi lý do)"
        edited_by = r.get("edited_by_name") or "?"

        with st.container(border=True):
            st.markdown(
                f"**{edited_by}** · {edited_at_str}  \n"
                f"📝 *{reason}*"
            )

            before = r.get("snapshot_before") or {}
            after = r.get("snapshot_after") or {}

            if fields:
                rows_diff = []
                for f in fields:
                    old_v = before.get(f)
                    new_v = after.get(f)
                    rows_diff.append({
                        "Field": f,
                        "Trước": "—" if old_v is None else str(old_v),
                        "Sau": "—" if new_v is None else str(new_v),
                    })
                st.dataframe(
                    pd.DataFrame(rows_diff),
                    hide_index=True,
                    use_container_width=True,
                )
            else:
                st.caption("(không có field nào đổi — chỉ có note)")


# ============================================================
# 💰 TÍNH LƯƠNG (Phase 6) — admin only CRUD, ke_toan xem
# ============================================================

_ADJUSTMENT_TYPE_LABEL = {
    "bonus_holiday":  "🎁 Thưởng lễ",
    "allowance_meal": "🍱 Phụ cấp ăn",
    "penalty":        "⚠️ Phạt",
    "other":          "📝 Khác",
}

_ITEM_TYPE_LABEL = {
    "shift":         "🕐 Ca làm",
    "monthly_fixed": "📅 Cố định tháng",
    "leave_paid":    "🌴 Nghỉ phép",
}


def _render_payroll():
    """Tab Tính lương: 3 sub-tabs — admin full, ke_toan xem."""
    if not is_ke_toan_or_admin():
        st.error("⛔ Chỉ admin và kế toán.")
        return

    sub = st.tabs([
        "📅 Kỳ lương",
        "💰 Bảng lương",
        "🎁 Phụ cấp / Thưởng",
    ])
    with sub[0]:
        _payroll_periods()
    with sub[1]:
        _payroll_bang_luong()
    with sub[2]:
        _payroll_adjustments()


def _load_payroll_periods() -> list[dict]:
    """List tất cả kỳ lương + items_count + total cho UI."""
    res = supabase.table("attendance_payroll_periods").select("*") \
        .order("start_date", desc=True).execute()
    periods = res.data or []
    if not periods:
        return []

    # Aggregate items count + total per period (single query)
    pids = [p["id"] for p in periods]
    items_res = supabase.table("attendance_payroll_items") \
        .select("period_id, salary_amount") \
        .in_("period_id", pids).execute()
    by_period_count: dict[int, int] = {}
    by_period_total: dict[int, int] = {}
    for r in (items_res.data or []):
        pid = r["period_id"]
        by_period_count[pid] = by_period_count.get(pid, 0) + 1
        by_period_total[pid] = by_period_total.get(pid, 0) + int(r.get("salary_amount") or 0)

    for p in periods:
        p["items_count"] = by_period_count.get(p["id"], 0)
        p["total_amount"] = by_period_total.get(p["id"], 0)
    return periods


def _payroll_periods():
    """Sub-tab Kỳ lương — CRUD + compute + finalize."""
    is_admin_user = is_admin()

    if is_admin_user:
        with st.expander("➕ Tạo kỳ lương mới", expanded=False):
            today = datetime.now(VN_TZ).date()
            first = today.replace(day=1)
            col_label, col_from, col_to = st.columns([3, 2, 2])
            with col_label:
                new_label = st.text_input(
                    "Tên kỳ", placeholder=f"vd: Tháng {today.month}/{today.year}",
                    key="cc_pr_new_label",
                )
            with col_from:
                new_from = st.date_input(
                    "Từ ngày", value=first, key="cc_pr_new_from"
                )
            with col_to:
                new_to = st.date_input(
                    "Đến ngày", value=today, key="cc_pr_new_to"
                )
            if st.button("💾 Tạo kỳ", type="primary",
                         disabled=not new_label.strip() or new_from > new_to,
                         key="cc_pr_new_save"):
                user = get_user()
                try:
                    supabase.table("attendance_payroll_periods").insert({
                        "label": new_label.strip(),
                        "start_date": new_from.isoformat(),
                        "end_date": new_to.isoformat(),
                        "created_by": user["id"],
                    }).execute()
                    log_action("ATT_PAYROLL_PERIOD_CREATE",
                               f"{new_label}: {new_from} → {new_to}")
                    st.toast(f"✓ Đã tạo kỳ '{new_label}'", icon="✅")
                    st.rerun()
                except Exception as e:
                    st.error(f"Lỗi: {e}")
    else:
        st.caption("ℹ️ Chỉ admin được tạo/tính/chốt kỳ. Bạn xem read-only.")

    periods = _load_payroll_periods()
    if not periods:
        st.info("📭 Chưa có kỳ lương nào. Admin tạo kỳ trước.")
        return

    st.markdown("---")
    for p in periods:
        pid = p["id"]
        is_finalized = p.get("status") == "finalized"
        status_badge = "🔒 Đã chốt" if is_finalized else "📂 Mở"
        items_count = p.get("items_count", 0)
        total = p.get("total_amount", 0)

        with st.container(border=True):
            col_info, col_meta, col_act = st.columns([3, 2, 2])
            with col_info:
                st.markdown(
                    f"**{p['label']}**  \n"
                    f"📅 {p['start_date']} → {p['end_date']}"
                )
            with col_meta:
                st.markdown(
                    f"{status_badge}  \n"
                    f"📋 {items_count} items · 💵 {total:,}đ"
                )
            with col_act:
                if is_admin_user and not is_finalized:
                    if st.button("💵 Tính lương",
                                 key=f"cc_pr_compute_{pid}",
                                 use_container_width=True,
                                 type="primary"):
                        _compute_period(pid)
                    if items_count > 0 and st.button(
                        "🔒 Chốt kỳ",
                        key=f"cc_pr_finalize_btn_{pid}",
                        use_container_width=True,
                    ):
                        _finalize_period_dialog(p)
                elif is_finalized:
                    finalized_at = p.get("finalized_at") or "?"
                    try:
                        ft = datetime.fromisoformat(finalized_at).astimezone(VN_TZ)
                        finalized_str = ft.strftime("%d/%m/%Y %H:%M")
                    except Exception:
                        finalized_str = finalized_at
                    st.caption(f"Chốt: {finalized_str}")


def _compute_period(period_id: int):
    """Trigger RPC compute_payroll_period."""
    try:
        res = call_rpc("compute_payroll_period", {"p_period_id": period_id})
        if isinstance(res, dict) and res.get("ok"):
            n = res.get("items_count", 0)
            total = res.get("total_amount", 0)
            st.toast(
                f"✓ Đã tính: {n} items · {total:,}đ",
                icon="✅",
            )
            st.rerun()
        else:
            err = (res or {}).get("error", "Lỗi RPC")
            st.error(f"❌ {err}")
    except Exception as e:
        st.error(f"Lỗi: {e}")


@st.dialog("🔒 Chốt kỳ lương")
def _finalize_period_dialog(period: dict):
    """Confirm dialog before finalize."""
    pid = period["id"]
    st.warning(
        f"⚠️ Sau khi chốt, kỳ lương **{period['label']}** sẽ **không tính lại được**. "
        "Mọi adjustments/items sẽ bị lock."
    )
    st.markdown(
        f"**Kỳ:** {period['label']}  \n"
        f"**Range:** {period['start_date']} → {period['end_date']}  \n"
        f"**Items:** {period.get('items_count', 0)}  \n"
        f"**Tổng:** {period.get('total_amount', 0):,}đ"
    )
    confirm = st.text_input(
        "Gõ `XÁC NHẬN` để chốt",
        key=f"cc_pr_fin_confirm_{pid}",
        placeholder="XÁC NHẬN",
    )
    can_finalize = confirm.strip().upper() == "XÁC NHẬN"

    if st.button("🔒 Chốt kỳ", type="primary", use_container_width=True,
                 disabled=not can_finalize,
                 key=f"cc_pr_fin_save_{pid}"):
        try:
            user = get_user()
            res = call_rpc("finalize_payroll_period", {
                "p_period_id": pid,
                "p_admin_id": user["id"],
            })
            if isinstance(res, dict) and res.get("ok"):
                st.toast(f"✓ Đã chốt kỳ '{period['label']}'", icon="🔒")
                st.rerun()
            else:
                err = (res or {}).get("error", "Lỗi RPC")
                st.error(f"❌ {err}")
        except Exception as e:
            st.error(f"Lỗi: {e}")


def _payroll_bang_luong():
    """Sub-tab Bảng lương — chọn period, xem breakdown per NV."""
    periods = _load_payroll_periods()
    if not periods:
        st.info("📭 Chưa có kỳ lương nào. Admin tạo + tính trước.")
        return

    options = {
        f"{p['label']} ({p['start_date']} → {p['end_date']}) — "
        f"{'🔒' if p.get('status') == 'finalized' else '📂'}": p
        for p in periods
    }
    picked_label = st.selectbox(
        "Chọn kỳ lương", list(options.keys()), key="cc_pr_bl_pick"
    )
    period = options[picked_label]
    pid = period["id"]

    if period.get("items_count", 0) == 0:
        st.warning(
            f"⚠️ Kỳ '{period['label']}' chưa được tính lương. "
            "Quay lại tab 'Kỳ lương' bấm '💵 Tính lương'."
        )
        return

    # Role filter
    visible_nv_ids = _visible_nv_ids()

    # Load items + adjustments
    q_items = supabase.table("attendance_payroll_items").select("*").eq("period_id", pid)
    if visible_nv_ids is not None:
        if not visible_nv_ids:
            st.info("Không có NV nào thuộc quyền xem.")
            return
        q_items = q_items.in_("nhan_vien_id", visible_nv_ids)
    items = q_items.execute().data or []

    q_adj = supabase.table("attendance_adjustments").select("*").eq("period_id", pid)
    if visible_nv_ids is not None:
        q_adj = q_adj.in_("nhan_vien_id", visible_nv_ids)
    adjustments = q_adj.execute().data or []

    if not items:
        st.info("Không có dữ liệu lương cho quyền xem hiện tại.")
        return

    # Load NV info
    nv_ids = list({i["nhan_vien_id"] for i in items} | {a["nhan_vien_id"] for a in adjustments})
    nv_res = supabase.table("nhan_vien").select("id, ho_ten, role") \
        .in_("id", nv_ids).execute()
    nv_map = {n["id"]: n for n in (nv_res.data or [])}

    # Aggregate per NV
    rows = []
    nv_summary: dict[int, dict] = {}
    for it in items:
        nv_id = it["nhan_vien_id"]
        if nv_id not in nv_summary:
            nv_summary[nv_id] = {
                "shift_minutes": 0, "ot_minutes": 0, "leave_minutes": 0,
                "luong_ca": 0, "rate": 0, "salary_type": "",
            }
        s = nv_summary[nv_id]
        s["luong_ca"] += int(it.get("salary_amount") or 0)
        s["rate"] = int(it.get("rate_snapshot") or 0)
        if it.get("item_type") == "shift":
            s["shift_minutes"] += int(it.get("worked_minutes") or 0)
            s["ot_minutes"] += int(it.get("ot_minutes") or 0)
            s["salary_type"] = "hourly"
        elif it.get("item_type") == "leave_paid":
            s["leave_minutes"] += int(it.get("worked_minutes") or 0)
            s["salary_type"] = "hourly"
        elif it.get("item_type") == "monthly_fixed":
            s["salary_type"] = "monthly_fixed"

    adj_sum: dict[int, int] = {}
    for a in adjustments:
        nv_id = a["nhan_vien_id"]
        adj_sum[nv_id] = adj_sum.get(nv_id, 0) + int(a.get("amount") or 0)

    for nv_id, s in nv_summary.items():
        nv = nv_map.get(nv_id, {})
        adj = adj_sum.get(nv_id, 0)
        total = s["luong_ca"] + adj
        salary_type_lbl = "⏰ Theo giờ" if s["salary_type"] == "hourly" else "📅 Cố định"
        rate_str = (f"{s['rate']:,}đ/giờ" if s["salary_type"] == "hourly"
                    else f"{s['rate']:,}đ/tháng")
        rows.append({
            "NV": nv.get("ho_ten", "?"),
            "Loại": salary_type_lbl,
            "Giờ làm": _format_minutes(s["shift_minutes"]) if s["salary_type"] == "hourly" else "—",
            "Giờ OT": _format_minutes(s["ot_minutes"]) if s["ot_minutes"] else "—",
            "Nghỉ phép": _format_minutes(s["leave_minutes"]) if s["leave_minutes"] else "—",
            "Đơn giá": rate_str,
            "Lương ca": f"{s['luong_ca']:,}đ",
            "Phụ cấp / Phạt": f"{adj:+,}đ" if adj else "—",
            "**Tổng cộng**": f"**{total:,}đ**",
            "_total": total,
        })

    if not rows:
        st.info("Không có data.")
        return

    df = pd.DataFrame(rows).sort_values("NV").reset_index(drop=True)
    display_cols = ["NV", "Loại", "Giờ làm", "Giờ OT", "Nghỉ phép",
                    "Đơn giá", "Lương ca", "Phụ cấp / Phạt", "**Tổng cộng**"]
    st.dataframe(df[display_cols], hide_index=True, use_container_width=True)

    # Overall metric
    grand_total = int(df["_total"].sum())
    st.metric("💵 Tổng lương kỳ này", f"{grand_total:,}đ")

    # 📥 Export Excel (Phase 7) — admin only
    if is_admin():
        try:
            xlsx_bytes = _export_payroll_excel(period)
            file_label = (
                (period.get("label") or f"period_{pid}")
                .replace("/", "-").replace(" ", "_")
            )
            file_date = datetime.now(VN_TZ).strftime("%Y%m%d")
            st.download_button(
                "📥 Export Excel (Bảng công + Bảng lương)",
                data=xlsx_bytes,
                file_name=f"bang_luong_{file_label}_{file_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                key=f"cc_pr_export_{pid}",
            )
        except Exception as e:
            st.warning(f"⚠️ Không tạo được file Excel: {e}")


def _payroll_adjustments():
    """Sub-tab Phụ cấp / Thưởng — CRUD adjustments per period."""
    is_admin_user = is_admin()
    periods = _load_payroll_periods()
    if not periods:
        st.info("📭 Chưa có kỳ lương nào.")
        return

    options = {
        f"{p['label']} ({p['start_date']} → {p['end_date']}) — "
        f"{'🔒' if p.get('status') == 'finalized' else '📂'}": p
        for p in periods
    }
    picked_label = st.selectbox(
        "Chọn kỳ lương", list(options.keys()), key="cc_pr_adj_pick"
    )
    period = options[picked_label]
    pid = period["id"]
    is_finalized = period.get("status") == "finalized"

    if is_admin_user and not is_finalized:
        with st.expander("➕ Thêm phụ cấp/thưởng", expanded=False):
            nv_list = load_all_nhan_vien(include_inactive=False)
            nv_options = {f"{n['ho_ten']} ({n.get('role','')})": n["id"]
                          for n in nv_list}
            col_nv, col_type = st.columns([3, 2])
            with col_nv:
                nv_lbl = st.selectbox(
                    "Nhân viên", list(nv_options.keys()),
                    key="cc_pr_adj_new_nv",
                )
            with col_type:
                adj_type = st.selectbox(
                    "Loại",
                    list(_ADJUSTMENT_TYPE_LABEL.keys()),
                    format_func=lambda k: _ADJUSTMENT_TYPE_LABEL[k],
                    key="cc_pr_adj_new_type",
                )
            amount = st.number_input(
                "Số tiền (đ) — âm = trừ lương (vd: phạt)",
                value=0, step=10000,
                key="cc_pr_adj_new_amount",
            )
            note = st.text_input(
                "Ghi chú", key="cc_pr_adj_new_note",
                placeholder="vd: Thưởng Tết 2026",
            )
            if st.button("💾 Thêm",
                         type="primary",
                         disabled=(amount == 0),
                         key="cc_pr_adj_new_save"):
                try:
                    user = get_user()
                    supabase.table("attendance_adjustments").insert({
                        "period_id": pid,
                        "nhan_vien_id": nv_options[nv_lbl],
                        "adjustment_type": adj_type,
                        "amount": int(amount),
                        "note": note or None,
                        "created_by": user["id"],
                    }).execute()
                    log_action("ATT_PAYROLL_ADJUSTMENT_ADD",
                               f"period={pid} nv={nv_options[nv_lbl]} "
                               f"type={adj_type} amount={amount}")
                    st.toast("✓ Đã thêm adjustment", icon="✅")
                    st.rerun()
                except Exception as e:
                    st.error(f"Lỗi: {e}")
    elif is_finalized:
        st.caption("🔒 Kỳ đã chốt — không thêm/xóa adjustments.")
    else:
        st.caption("ℹ️ Chỉ admin được thêm/xóa adjustments.")

    # List adjustments (role-aware)
    visible_nv_ids = _visible_nv_ids()
    q = supabase.table("attendance_adjustments").select("*").eq("period_id", pid)
    if visible_nv_ids is not None:
        if not visible_nv_ids:
            st.info("Không có adjustment nào thuộc quyền xem.")
            return
        q = q.in_("nhan_vien_id", visible_nv_ids)
    rows = q.order("created_at", desc=True).execute().data or []

    if not rows:
        st.info("📭 Chưa có phụ cấp/thưởng nào cho kỳ này.")
        return

    nv_ids = list({a["nhan_vien_id"] for a in rows})
    nv_res = supabase.table("nhan_vien").select("id, ho_ten").in_("id", nv_ids).execute()
    nv_map = {n["id"]: n["ho_ten"] for n in (nv_res.data or [])}

    for r in rows:
        adj_id = r["id"]
        type_lbl = _ADJUSTMENT_TYPE_LABEL.get(
            r["adjustment_type"], r["adjustment_type"]
        )
        nv_name = nv_map.get(r["nhan_vien_id"], "?")
        amount = int(r.get("amount") or 0)
        sign = "+" if amount > 0 else ""
        with st.container(border=True):
            col_info, col_amount, col_del = st.columns([4, 2, 1])
            with col_info:
                st.markdown(
                    f"**{nv_name}** · {type_lbl}  \n"
                    f"📝 {r.get('note') or '(không note)'}"
                )
            with col_amount:
                color = "#1a7f37" if amount > 0 else "#cf4c2c"
                st.markdown(
                    f"<div style='font-size:1.1rem;font-weight:700;color:{color};'>"
                    f"{sign}{amount:,}đ</div>",
                    unsafe_allow_html=True,
                )
            with col_del:
                if is_admin_user and not is_finalized:
                    if st.button("🗑", key=f"cc_pr_adj_del_{adj_id}",
                                 use_container_width=True,
                                 help="Xóa adjustment"):
                        try:
                            supabase.table("attendance_adjustments") \
                                .delete().eq("id", adj_id).execute()
                            log_action("ATT_PAYROLL_ADJUSTMENT_DELETE",
                                       f"id={adj_id}")
                            st.toast("🗑 Đã xóa", icon="🗑")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Lỗi: {e}")


# ============================================================
# 📥 EXPORT EXCEL (Phase 7) — admin only, 2 sheets per period
# ============================================================

def _export_payroll_excel(period: dict) -> bytes:
    """Generate XLSX file (2 sheets: Bảng công + Bảng lương) cho 1 kỳ lương.

    Sheet 1 "Bảng công": session rows (NV, ngày, ca, vào, ra, worked, late, status)
    Sheet 2 "Bảng lương": payroll breakdown per NV (giờ, OT, rate, lương, adj, tổng)

    Format A4 landscape, fit-to-width, header bold + bg xám.
    Footer: "Lập bởi: <admin> · <date>".
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins, PrintOptions
    from io import BytesIO

    pid = period["id"]
    period_label = period.get("label", f"Period {pid}")
    user = get_user() or {}
    admin_name = user.get("ho_ten", "?")
    today_str = datetime.now(VN_TZ).strftime("%d/%m/%Y %H:%M")

    # Date range
    try:
        date_from = datetime.strptime(period["start_date"], "%Y-%m-%d").date()
        date_to = datetime.strptime(period["end_date"], "%Y-%m-%d").date()
    except Exception:
        date_from = date.today()
        date_to = date.today()

    # Common styles
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=11, color="1A1A2E")
    header_fill = PatternFill("solid", fgColor="EEEEEE")
    title_font = Font(bold=True, size=14, color="E63946")
    footer_font = Font(italic=True, size=9, color="888888")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    money_fmt = '#,##0"đ"'

    wb = Workbook()

    # ─────────────────────────────────────────────────────────
    # Sheet 1: Bảng công
    # ─────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Bảng công"
    ws1.page_setup.orientation = ws1.ORIENTATION_LANDSCAPE
    ws1.page_setup.paperSize = ws1.PAPERSIZE_A4
    ws1.page_setup.fitToWidth = 1
    ws1.page_setup.fitToHeight = 0
    ws1.sheet_properties.pageSetUpPr.fitToPage = True
    ws1.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws1.print_options.horizontalCentered = True

    ws1.cell(row=1, column=1, value=f"BẢNG CÔNG — {period_label}").font = title_font
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws1.cell(row=1, column=1).alignment = center

    ws1.cell(row=2, column=1,
             value=f"Kỳ: {date_from.strftime('%d/%m/%Y')} → {date_to.strftime('%d/%m/%Y')}")
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws1.cell(row=2, column=1).alignment = center

    sheet1_headers = ["STT", "Ngày", "Thứ", "NV", "Ca", "Chi nhánh",
                      "Vào", "Ra", "Worked (phút)", "Late (phút)", "Status"]
    header_row = 4
    for col, h in enumerate(sheet1_headers, start=1):
        c = ws1.cell(row=header_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center

    # Load sessions (no role filter for export — admin xem all)
    sessions = _load_sessions_for_range(date_from, date_to, branch=None, nv_id_filter=None)

    row = header_row + 1
    for idx, s in enumerate(sessions, start=1):
        try:
            wd = datetime.strptime(s["work_date"], "%Y-%m-%d").date()
            ngay_str = wd.strftime("%d/%m/%Y")
            dow = DOW_VN[wd.weekday()]
        except Exception:
            ngay_str = s.get("work_date", "")
            dow = ""

        check_in = _format_time_vn(s.get("check_in_at"))
        check_out = _format_time_vn(s.get("check_out_at"))

        status = s.get("status", "open")
        status_lbl = _STATUS_LABEL_VN.get(status, status)
        if s.get("is_auto_checkout") and status == "auto_closed":
            status_lbl += " ⚠"

        worked = int(s.get("worked_minutes", 0) or 0)
        late = int(s.get("late_minutes", 0) or 0) if s.get("is_late") else 0

        values = [
            idx, ngay_str, dow, s.get("ho_ten", "?"),
            s.get("shift_label", "?"), s.get("branch_name", ""),
            check_in, check_out,
            worked, late, status_lbl,
        ]
        for col, v in enumerate(values, start=1):
            c = ws1.cell(row=row, column=col, value=v)
            c.border = border
            c.alignment = left if col in (4, 5, 6, 11) else center
        row += 1

    # Total worked row
    if sessions:
        total_worked = sum(int(s.get("worked_minutes", 0) or 0) for s in sessions)
        c = ws1.cell(row=row, column=8, value="Tổng worked:")
        c.font = Font(bold=True)
        c.alignment = right
        c = ws1.cell(row=row, column=9, value=total_worked)
        c.font = Font(bold=True)
        c.border = border
        c.alignment = center
        row += 1

    # Footer
    row += 1
    ws1.cell(row=row, column=1,
             value=f"Lập bởi: {admin_name} · {today_str}").font = footer_font
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)

    # Column widths
    col_widths_1 = [5, 12, 6, 22, 18, 18, 8, 8, 14, 12, 22]
    for i, w in enumerate(col_widths_1, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ─────────────────────────────────────────────────────────
    # Sheet 2: Bảng lương
    # ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Bảng lương")
    ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE
    ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
    ws2.page_setup.fitToWidth = 1
    ws2.page_setup.fitToHeight = 0
    ws2.sheet_properties.pageSetUpPr.fitToPage = True
    ws2.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws2.print_options.horizontalCentered = True

    ws2.cell(row=1, column=1, value=f"BẢNG LƯƠNG — {period_label}").font = title_font
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws2.cell(row=1, column=1).alignment = center

    finalized_str = ""
    if period.get("status") == "finalized" and period.get("finalized_at"):
        try:
            ft = datetime.fromisoformat(period["finalized_at"]).astimezone(VN_TZ)
            finalized_str = f"🔒 Chốt: {ft.strftime('%d/%m/%Y %H:%M')}"
        except Exception:
            pass
    sub_meta = f"Kỳ: {date_from.strftime('%d/%m/%Y')} → {date_to.strftime('%d/%m/%Y')}"
    if finalized_str:
        sub_meta += " · " + finalized_str
    ws2.cell(row=2, column=1, value=sub_meta)
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws2.cell(row=2, column=1).alignment = center

    sheet2_headers = ["STT", "Nhân viên", "Loại lương",
                      "Giờ regular", "Giờ OT", "Nghỉ phép (giờ)",
                      "Đơn giá (đ)", "Lương ca (đ)",
                      "Phụ cấp / Phạt (đ)", "Tổng cộng (đ)"]
    for col, h in enumerate(sheet2_headers, start=1):
        c = ws2.cell(row=header_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center

    # Load items + adjustments (admin = all NV)
    items = supabase.table("attendance_payroll_items").select("*") \
        .eq("period_id", pid).execute().data or []
    adjustments = supabase.table("attendance_adjustments").select("*") \
        .eq("period_id", pid).execute().data or []

    nv_ids = list({i["nhan_vien_id"] for i in items}
                  | {a["nhan_vien_id"] for a in adjustments})
    nv_map: dict[int, dict] = {}
    if nv_ids:
        nv_res = supabase.table("nhan_vien").select("id, ho_ten") \
            .in_("id", nv_ids).execute()
        nv_map = {n["id"]: n for n in (nv_res.data or [])}

    # Aggregate per NV
    nv_summary: dict[int, dict] = {}
    for it in items:
        nv_id = it["nhan_vien_id"]
        if nv_id not in nv_summary:
            nv_summary[nv_id] = {
                "shift_minutes": 0, "ot_minutes": 0, "leave_minutes": 0,
                "luong_ca": 0, "rate": 0, "salary_type": "",
            }
        s = nv_summary[nv_id]
        s["luong_ca"] += int(it.get("salary_amount") or 0)
        s["rate"] = int(it.get("rate_snapshot") or 0)
        if it.get("item_type") == "shift":
            s["shift_minutes"] += int(it.get("worked_minutes") or 0)
            s["ot_minutes"] += int(it.get("ot_minutes") or 0)
            s["salary_type"] = "hourly"
        elif it.get("item_type") == "leave_paid":
            s["leave_minutes"] += int(it.get("worked_minutes") or 0)
            s["salary_type"] = "hourly"
        elif it.get("item_type") == "monthly_fixed":
            s["salary_type"] = "monthly_fixed"

    adj_sum: dict[int, int] = {}
    for a in adjustments:
        nv_id = a["nhan_vien_id"]
        adj_sum[nv_id] = adj_sum.get(nv_id, 0) + int(a.get("amount") or 0)

    row = header_row + 1
    grand_total = 0
    sorted_nv_ids = sorted(nv_summary.keys(),
                            key=lambda nid: nv_map.get(nid, {}).get("ho_ten", ""))
    for idx, nv_id in enumerate(sorted_nv_ids, start=1):
        s = nv_summary[nv_id]
        nv = nv_map.get(nv_id, {})
        adj = adj_sum.get(nv_id, 0)
        total = s["luong_ca"] + adj
        grand_total += total

        salary_type_lbl = ("Theo giờ" if s["salary_type"] == "hourly"
                           else "Cố định tháng")
        gio_regular = round(s["shift_minutes"] / 60.0, 2) if s["salary_type"] == "hourly" else None
        gio_ot = round(s["ot_minutes"] / 60.0, 2) if s["ot_minutes"] else None
        gio_leave = round(s["leave_minutes"] / 60.0, 2) if s["leave_minutes"] else None

        values_2 = [
            idx, nv.get("ho_ten", "?"), salary_type_lbl,
            gio_regular, gio_ot, gio_leave,
            s["rate"], s["luong_ca"], adj if adj else None, total,
        ]
        for col, v in enumerate(values_2, start=1):
            c = ws2.cell(row=row, column=col, value=v)
            c.border = border
            if col in (1,):
                c.alignment = center
            elif col in (2, 3):
                c.alignment = left
            else:
                c.alignment = right
            # Number format cho money cols (7-10) + hour cols (4-6)
            if col in (7, 8, 9, 10):
                c.number_format = money_fmt
            elif col in (4, 5, 6) and v is not None:
                c.number_format = "0.00"
        row += 1

    # Grand total row
    if sorted_nv_ids:
        c = ws2.cell(row=row, column=9, value="TỔNG CỘNG:")
        c.font = Font(bold=True)
        c.alignment = right
        c = ws2.cell(row=row, column=10, value=grand_total)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFF0F1")
        c.border = border
        c.alignment = right
        c.number_format = money_fmt
        row += 1

    # Adjustments breakdown
    if adjustments:
        row += 1
        c = ws2.cell(row=row, column=1, value="CHI TIẾT PHỤ CẤP / PHẠT")
        c.font = Font(bold=True, size=11)
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 1
        adj_headers = ["NV", "Loại", "Số tiền", "Ghi chú"]
        for col, h in enumerate(adj_headers, start=1):
            c = ws2.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = border
            c.alignment = center
        row += 1
        for a in adjustments:
            adj_type = _ADJUSTMENT_TYPE_LABEL.get(
                a.get("adjustment_type", ""), a.get("adjustment_type", "")
            )
            nv_name = nv_map.get(a["nhan_vien_id"], {}).get("ho_ten", "?")
            amt = int(a.get("amount") or 0)
            note = a.get("note") or ""
            for col, v in enumerate([nv_name, adj_type, amt, note], start=1):
                c = ws2.cell(row=row, column=col, value=v)
                c.border = border
                if col == 3:
                    c.alignment = right
                    c.number_format = money_fmt
                else:
                    c.alignment = left
            row += 1

    # Footer
    row += 1
    ws2.cell(row=row, column=1,
             value=f"Lập bởi: {admin_name} · {today_str}").font = footer_font
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    # Column widths
    col_widths_2 = [5, 22, 14, 12, 10, 14, 14, 14, 16, 16]
    for i, w in enumerate(col_widths_2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Output bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
